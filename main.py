import threading
import xlrd
import time
import requests
import openpyxl
from bs4 import BeautifulSoup as BS
import excel2img
from PIL import Image
import xlwt
import os
import pyexcel as p
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import telebot
import sqlite3
from telebot import types
from openpyxl.styles import PatternFill, Border, Side

URL = 'http://students.perm.hse.ru/timetable/?ysclid=lh8xpt9dhk529019377'
API_KEY = '6171088054:AAEsxn8wde9NAvs_vYGC2FJXTRvJzfT-iNk'

response = requests.get(URL)
soup = BS(response.content, 'html.parser')
print(soup.title.string)


# создание файлов с расписанием
def save_from_www(link, filename):
    print(filename)
    r = requests.get(link, allow_redirects=True)
    open(filename, "wb").write(r.content)


# заполняем массив ссылками на файлы с расписанием
internalLinks = []
shedule_names = []
i = 1
for a in soup.find_all('a'):
    if ("Расписание занятий (") in str(a.text) and i <= 2:
        i += 1
        shedule_names.append(str(a.text))
        internalLinks.append(a.get('href'))
number_of_schedules = len(internalLinks)
# корректируем ссылки
if number_of_schedules != 0:
    for i in range(number_of_schedules):
        shedule_names[i] = shedule_names[i].replace("Расписание занятий ", "")
    for i in range(number_of_schedules):
        if ('//www.hse.ru') in internalLinks[i]:
            internalLinks[i] = internalLinks[i].replace('//www.hse.ru', 'http://students.perm.hse.ru')
        elif ('www.hse.ru') in internalLinks[i]:
            internalLinks[i] = internalLinks[i].replace('www.hse.ru', 'http://students.perm.hse.ru')
        elif ('http://students.perm.hse.ru') not in internalLinks[i]:
            internalLinks[i] = 'http://students.perm.hse.ru' + internalLinks[i]
    print(internalLinks)
    print(shedule_names)
print(number_of_schedules)

# создаем файлы с расписанием
if number_of_schedules == 1:
    save_from_www(internalLinks[0], 'this_week.xls')
    this_week = xlrd.open_workbook('this_week.xls', formatting_info=True)
elif number_of_schedules == 2:
    save_from_www(internalLinks[0], 'this_week.xls')
    save_from_www(internalLinks[1], 'next_week.xls')
    this_week = xlrd.open_workbook('this_week.xls', formatting_info=True)
    next_week = xlrd.open_workbook('next_week.xls', formatting_info=True)

''' 
Работа с файлом 
1 - разъединяем ячейки
2 - конвертируем в xlsx
3 - работаем с ячейками'''


# разъединяем ячейки
def unmerged_cell(name, file):
    excel = xlwt.Workbook()
    for rd_sheet in file.sheets():
        # for each sheet
        wt_sheet = excel.add_sheet(rd_sheet.name)

        writed_cells = []

        # overwrite for merged cells
        for crange in rd_sheet.merged_cells:
            # for each merged_cell
            rlo, rhi, clo, chi = crange
            cell_value = rd_sheet.cell(rlo, clo).value
            for rowx in range(rlo, rhi):
                for colx in range(clo, chi):
                    wt_sheet.write(rowx, colx, cell_value)
                    writed_cells.append((rowx, colx))

        # write all un-merged cells
        for r in range(0, rd_sheet.nrows):
            for c in range(0, rd_sheet.ncols):
                if (r, c) in writed_cells:
                    continue
                cell_value = rd_sheet.cell(r, c).value
                wt_sheet.write(r, c, cell_value)

        # save the un-merged excel file
    (origin_file, ext) = os.path.splitext(name)
    unmerge_excel_file = origin_file + ext
    excel.save(unmerge_excel_file)


# разъединяем ячейки
if number_of_schedules == 1:
    unmerged_cell('this_week.xls', this_week)
elif number_of_schedules == 2:
    unmerged_cell('this_week.xls', this_week)
    unmerged_cell('next_week.xls', next_week)


# конвертируем в xlsx
def convert_to_xlxs(file_name):
    p.save_book_as(file_name=file_name + '.xls',
                   dest_file_name=file_name + '.xlsx')


# конвертируем в xlsx
if number_of_schedules == 1:
    convert_to_xlxs('this_week')
    this_week = load_workbook('this_week.xlsx')
    this_week_courses = [this_week["1 курс"], this_week["2 курс"], this_week["3 курс"], this_week["4 курс"]]
elif number_of_schedules == 2:
    convert_to_xlxs('this_week')
    convert_to_xlxs('next_week')
    this_week = load_workbook('this_week.xlsx')
    next_week = load_workbook('next_week.xlsx')
    this_week_courses = [this_week["1 курс"], this_week["2 курс"], this_week["3 курс"], this_week["4 курс"]]
    next_week_courses = [next_week["1 курс"], next_week["2 курс"], next_week["3 курс"], next_week["4 курс"]]

letters = ["A", 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
           'W']


# работа с ячейками
def cells(course_sheet, file, name):
    for i in letters:
        course_sheet.column_dimensions[i].width = 30
        for j in range(1, 40):
            course_sheet[i + str(j)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    file.save(name)
    course_sheet.column_dimensions['A'].width = 15
    course_sheet.column_dimensions['B'].width = 15


# работа с ячейками
if number_of_schedules == 1:
    convert_to_xlxs('this_week')
    this_week = load_workbook('this_week.xlsx')
    this_week_courses = [this_week["1 курс"], this_week["2 курс"], this_week["3 курс"], this_week["4 курс"]]
    for i in this_week_courses:
        cells(i, this_week, 'this_week.xlsx')
elif number_of_schedules == 2:
    convert_to_xlxs('this_week')
    convert_to_xlxs('next_week')
    this_week = load_workbook('this_week.xlsx')
    next_week = load_workbook('next_week.xlsx')
    this_week_courses = [this_week["1 курс"], this_week["2 курс"], this_week["3 курс"], this_week["4 курс"]]
    next_week_courses = [next_week["1 курс"], next_week["2 курс"], next_week["3 курс"], next_week["4 курс"]]
    for i in this_week_courses:
        cells(i, this_week, 'this_week.xlsx')
    for i in next_week_courses:
        cells(i, next_week, 'next_week.xlsx')

#цвета
thins = Side(border_style="thin", color="000000")
monday = PatternFill('solid', fgColor="CCFFFF")
tuesday = PatternFill('solid', fgColor="FFCCFF")
wednesday = PatternFill('solid', fgColor="CCFFCC")
thursday = PatternFill('solid', fgColor="FFFFCC")
friday = PatternFill('solid', fgColor="FFCCCC")
saturday = PatternFill('solid', fgColor="CCECFF")
#меняет цвета ячеек
def colour(filename):
    week = load_workbook(filename)
    courses = [week["1 курс"], week["2 курс"], week["3 курс"], week["4 курс"]]
    for cource in courses:
        for j in range(3, 40):
            for i in letters:
                if cource["A" + str(j)].value is not None and "Вторник" in cource["A" + str(j)].value:
                    cource[i + str(j)].fill = tuesday
                elif cource["A" + str(j)].value is not None and "Среда" in cource["A" + str(j)].value:
                    cource[i + str(j)].fill = wednesday
                elif cource["A" + str(j)].value is not None and "Понедельник" in cource["A" + str(j)].value:
                    cource[i + str(j)].fill = monday
                elif cource["A" + str(j)].value is not None and "Четверг" in cource["A" + str(j)].value:
                    cource[i + str(j)].fill = thursday
                elif cource["A" + str(j)].value is not None and "Пятница" in cource["A" + str(j)].value:
                    cource[i + str(j)].fill = friday
                elif cource["A" + str(j)].value is not None and "Суббота" in cource["A" + str(j)].value:
                    cource[i + str(j)].fill = saturday
                cource[i + str(j)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
    week.save(filename)

if number_of_schedules == 1:
    colour('this_week.xlsx')
elif number_of_schedules == 2:
    colour('this_week.xlsx')
    colour('next_week.xlsx')

# массивы с картинками
timetable_pics = []
this_week_pics = []
next_week_pics = []
timetable_pics.append(this_week_pics)
timetable_pics.append(next_week_pics)


# объединяет картинки
def merge_images(file_name1, file_name2, name_new):
    image_1 = Image.open(file_name1)
    image_2 = Image.open(file_name2)
    size_1 = image_1.size
    size_2 = image_2.size
    new_image = Image.new('RGB', (size_1[0] + size_2[0], size_1[1]), (250, 250, 250))
    new_image.paste(image_1, (0, 0))
    new_image.paste(image_2, (size_1[0], 0))
    new_image.save(name_new)
    # os.remove(file_name2)
    return new_image

# создаем картинки
def create_images(main_name, week):
    timetable_pics[week - 1] = []

    timetable = openpyxl.load_workbook(main_name)
    # узнаем кол-во рядов и столбцов на каждом листе
    first_course = timetable["1 курс"]
    number_1 = 0
    help = first_course.max_row + 3
    for i in range(3, help):
        inf = str(first_course[i][0].value)
        if inf == 'None':
            number_1 = i - 3
            break
    number_of_columns_1 = first_course.max_column

    second_course = timetable["2 курс"]
    number_2 = 0
    for i in range(3, help):
        inf = str(second_course[i][0].value)
        if inf == 'None':
            number_2 = i - 3
            break
    number_of_columns_2 = second_course.max_column

    third_course = timetable["3 курс"]
    number_3 = 0
    for i in range(3, help):
        inf = str(third_course[i][0].value)
        if inf == 'None':
            number_3 = i - 3
            break
    number_of_columns_3 = third_course.max_column

    fourth_course = timetable["4 курс"]
    number_4 = 0
    for i in range(3, help):
        inf = str(fourth_course[i][0].value)
        if inf == 'None':
            number_4 = i - 3
            break
    number_of_columns_4 = fourth_course.max_column

    # направления 1 курса
    ris_fin_1 = -1
    mb_fin_1 = -1
    i_fin_1 = -1
    u_fin_1 = -1
    ia_fin_1 = -1
    ris_start_1 = -1
    mb_start_1 = -1
    i_start_1 = -1
    u_start_1 = -1
    ia_start_1 = -1

    # направления 2 курса
    bi_start_2 = -1
    pi_start_2 = -1
    e_start_2 = -1
    ub_start_2 = -1
    u_start_2 = -1
    ia_start_2 = -1
    i_start_2 = -1
    bi_fin_2 = -1
    pi_fin_2 = -1
    e_fin_2 = -1
    ub_fin_2 = -1
    u_fin_2 = -1
    ia_fin_2 = -1
    i_fin_2 = -1

    # направления 3 курса
    bi_start_3 = -1
    pi_start_3 = -1
    e_start_3 = -1
    ub_start_3 = -1
    u_start_3 = -1
    i_start_3 = -1
    bi_fin_3 = -1
    pi_fin_3 = -1
    e_fin_3 = -1
    ub_fin_3 = -1
    u_fin_3 = -1
    i_fin_3 = -1

    # направления 4 курса
    i_start_4 = -1
    i_fin_4 = -1
    # находим диапазоны для каждого направления
    for i in range(number_of_columns_1):
        inf = str(first_course[3][i].value)  # inf = first_course.cell_value(2, i)
        if "РИС" in inf:
            if (ris_start_1 == -1):
                ris_start_1 = i
            ris_fin_1 = i
        elif "МБ" in inf:
            if (mb_start_1 == -1):
                mb_start_1 = i
            mb_fin_1 = i
        elif "ИЯ" in inf:
            if (ia_start_1 == -1):
                ia_start_1 = i
            ia_fin_1 = i
        elif "И" in inf:
            if (i_start_1 == -1):
                i_start_1 = i
            i_fin_1 = i
        elif "Ю" in inf:
            if (u_start_1 == -1):
                u_start_1 = i
            u_fin_1 = i

    for i in range(number_of_columns_2):
        inf = str(second_course[3][i].value)
        if "БИ" in inf:
            if (bi_start_2 == -1):
                bi_start_2 = i
            bi_fin_2 = i
        elif "ПИ" in inf:
            if (pi_start_2 == -1):
                pi_start_2 = i
            pi_fin_2 = i
        elif "Э" in inf:
            if (e_start_2 == -1):
                e_start_2 = i
            e_fin_2 = i
        elif "УБ" in inf:
            if (ub_start_2 == -1):
                ub_start_2 = i
            ub_fin_2 = i
        elif "Ю" in inf:
            if (u_start_2 == -1):
                u_start_2 = i
            u_fin_2 = i
        elif "ИЯ" in inf:
            if (ia_start_2 == -1):
                ia_start_2 = i
            ia_fin_2 = i
        elif "И" in inf:
            if (i_start_2 == -1):
                i_start_2 = i
            i_fin_2 = i

    for i in range(number_of_columns_3):
        inf = str(third_course[3][i].value)
        if "БИ" in inf:
            if (bi_start_3 == -1):
                bi_start_3 = i
            bi_fin_3 = i
        elif "ПИ" in inf:
            if (pi_start_3 == -1):
                pi_start_3 = i
            pi_fin_3 = i
        elif "Э" in inf:
            if (e_start_3 == -1):
                e_start_3 = i
            e_fin_3 = i
        elif "УБ" in inf:
            if (ub_start_3 == -1):
                ub_start_3 = i
            ub_fin_3 = i
        elif "Ю" in inf:
            if (u_start_3 == -1):
                u_start_3 = i
            u_fin_3 = i
        elif "И" in inf:
            if (i_start_3 == -1):
                i_start_3 = i
            i_fin_3 = i
    # красиво формируем диапазоны
    rows_time_1 = "A3:B" + str(number_1)
    if (ris_start_1 != -1):
        course_1_RIS_rows = letters[ris_start_1] + "3:" + letters[ris_fin_1] + str(number_1)
    else:
        course_1_RIS_rows = -1
    if (mb_start_1 != -1):
        course_1_MB_rows = letters[mb_start_1] + "3:" + letters[mb_fin_1] + str(number_1)
    else:
        course_1_MB_rows = -1
    if (i_start_1 != -1):
        course_1_I_rows = letters[i_start_1] + "3:" + letters[i_fin_1] + str(number_1)
    else:
        course_1_I_rows = -1
    if (u_start_1 != -1):
        course_1_U_rows = letters[u_start_1] + "3:" + letters[u_fin_1] + str(number_1)
    else:
        course_1_U_rows = -1
    if (ia_start_1 != -1):
        course_1_IA_rows = letters[ia_start_1] + "3:" + letters[ia_fin_1] + str(number_1)
    else:
        course_1_IA_rows = -1

    rows_time_2 = "A3:B" + str(number_2)
    if (bi_start_2 != -1):
        course_2_BI_rows = letters[bi_start_2] + "3:" + letters[bi_fin_2] + str(number_2)
    else:
        course_2_BI_rows = -1
    if (pi_start_2 != -1):
        course_2_PI_rows = letters[pi_start_2] + "3:" + letters[pi_fin_2] + str(number_2)
    else:
        course_2_PI_rows = -1
    if (e_start_2 != -1):
        course_2_E_rows = letters[e_start_2] + "3:" + letters[e_fin_2] + str(number_2)
    else:
        course_2_E_rows = -1
    if (ub_start_2 != -1):
        course_2_UB_rows = letters[ub_start_2] + "3:" + letters[ub_fin_2] + str(number_2)
    else:
        course_2_UB_rows = -1
    if (u_start_2 != -1):
        course_2_U_rows = letters[u_start_2] + "3:" + letters[u_fin_2] + str(number_2)
    else:
        course_2_U_rows = -1
    if (ia_start_2 != -1):
        course_2_IA_rows = letters[ia_start_2] + "3:" + letters[ia_fin_2] + str(number_2)
    else:
        course_2_IA_rows = -1
    if (i_start_2 != -1):
        course_2_I_rows = letters[i_start_2] + "3:" + letters[i_fin_2] + str(number_2)
    else:
        course_2_I_rows = -1

    rows_time_3 = "A3:B" + str(number_3)
    if (i_start_2 != -1):
        course_3_BI_rows = letters[bi_start_3] + "3:" + letters[bi_fin_3] + str(number_3)
    else:
        course_3_BI_rows = -1
    if (pi_start_3 != -1):
        course_3_PI_rows = letters[pi_start_3] + "3:" + letters[pi_fin_3] + str(number_3)
    else:
        course_3_PI_rows = -1
    if (e_start_3 != -1):
        course_3_E_rows = letters[e_start_3] + "3:" + letters[e_fin_3] + str(number_3)
    else:
        course_3_E_rows = -1
    if (ub_start_3 != -1):
        course_3_UB_rows = letters[ub_start_3] + "3:" + letters[ub_fin_3] + str(number_3)
    else:
        course_3_UB_rows = -1
    if (u_start_3 != -1):
        course_3_U_rows = letters[u_start_3] + "3:" + letters[u_fin_3] + str(number_3)
    else:
        course_3_U_rows = -1
    if (i_start_3 != -1):
        course_3_I_rows = letters[i_start_3] + "3:" + letters[i_fin_3] + str(number_3)
    else:
        course_3_I_rows = -1

    rows_time_4 = "A3:B" + str(number_4)
    course_4_I_rows = "C3:C" + str(number_4)
    # создаем картинки с расписанием
    excel2img.export_img(main_name, str(week) + "main1.png", "1 курс", rows_time_1)
    if (course_1_RIS_rows != -1):
        excel2img.export_img(main_name, str(week) + "wRIS1.png", "1 курс", course_1_RIS_rows)
        new_image = merge_images(str(week) + "main1.png", str(week) + "wRIS1.png", str(week) + "wRIS1.png") # склеиваем картинки с парами и днями недели
        timetable_pics[week - 1].append(str(week) + "wRIS1.png") # заносим в массив
    else:
        timetable_pics[week - 1].append("empty")
    if (course_1_MB_rows != -1):
        excel2img.export_img(main_name, str(week) + "wMB1.png", "1 курс", course_1_MB_rows)
        new_image = merge_images(str(week) + "main1.png", str(week) + "wMB1.png", str(week) + "wMB1.png")
        timetable_pics[week - 1].append(str(week) + "wMB1.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_1_I_rows != -1):
        excel2img.export_img(main_name, str(week) + "wI1.png", "1 курс", course_1_I_rows)
        new_image = merge_images(str(week) + "main1.png", str(week) + "wI1.png", str(week) + "wI1.png")
        timetable_pics[week - 1].append(str(week) + "wI1.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_1_U_rows != -1):
        excel2img.export_img(main_name, str(week) + "wU1.png", "1 курс", course_1_U_rows)
        new_image = merge_images(str(week) + "main1.png", str(week) + "wU1.png", str(week) + "wU1.png")
        timetable_pics[week - 1].append(str(week) + "wU1.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_1_IA_rows != -1):
        excel2img.export_img(main_name, str(week) + "wIA1.png", "1 курс", course_1_IA_rows)
        new_image = merge_images(str(week) + "main1.png", str(week) + "wIA1.png", str(week) + "wIA1.png")
        timetable_pics[week - 1].append(str(week) + "wIA1.png")
    else:
        timetable_pics[week - 1].append("empty")
    excel2img.export_img(main_name, str(week) + "main2.png", "2 курс", rows_time_2)
    if (course_2_BI_rows != -1):
        excel2img.export_img(main_name, str(week) + "wBI2.png", "2 курс", course_2_BI_rows)
        new_image = merge_images(str(week) + "main2.png", str(week) + "wBI2.png", str(week) + "wBI2.png")
        timetable_pics[week - 1].append(str(week) + "wBI2.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_2_PI_rows != -1):
        excel2img.export_img(main_name, str(week) + "wPI2.png", "2 курс", course_2_PI_rows)
        new_image = merge_images(str(week) + "main2.png", str(week) + "wPI2.png", str(week) + "wPI2.png")
        timetable_pics[week - 1].append(str(week) + "wPI2.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_2_E_rows != -1):
        excel2img.export_img(main_name, str(week) + "wE2.png", "2 курс", course_2_E_rows)
        new_image = merge_images(str(week) + "main2.png", str(week) + "wE2.png", str(week) + "wE2.png")
        timetable_pics[week - 1].append(str(week) + "wE2.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_2_UB_rows != -1):
        excel2img.export_img(main_name, str(week) + "wUB2.png", "2 курс", course_2_UB_rows)
        new_image = merge_images(str(week) + "main2.png", str(week) + "wUB2.png", str(week) + "wUB2.png")
        timetable_pics[week - 1].append(str(week) + "wUB2.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_2_I_rows != -1):
        excel2img.export_img(main_name, str(week) + "wI2.png", "2 курс", course_2_I_rows)
        new_image = merge_images(str(week) + "main2.png", str(week) + "wI2.png", str(week) + "wI2.png")
        timetable_pics[week - 1].append(str(week) + "wI2.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_2_U_rows != -1):
        excel2img.export_img(main_name, str(week) + "wU2.png", "2 курс", course_2_U_rows)
        new_image = merge_images(str(week) + "main2.png", str(week) + "wU2.png", str(week) + "wU2.png")
        timetable_pics[week - 1].append(str(week) + "wU2.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_2_IA_rows != -1):
        excel2img.export_img(main_name, str(week) + "wIA2.png", "2 курс", course_2_IA_rows)
        new_image = merge_images(str(week) + "main2.png", str(week) + "wIA2.png", str(week) + "wIA2.png")
        timetable_pics[week - 1].append(str(week) + "wIA2.png")
    else:
        timetable_pics[week - 1].append("empty")
    excel2img.export_img(main_name, str(week) + "main3.png", "3 курс", rows_time_3)
    if (course_3_BI_rows != -1):
        excel2img.export_img(main_name, str(week) + "wBI3.png", "3 курс", course_3_BI_rows)
        new_image = merge_images(str(week) + "main3.png", str(week) + "wBI3.png", str(week) + "wBI3.png")
        timetable_pics[week - 1].append(str(week) + "wBI3.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_3_PI_rows != -1):
        excel2img.export_img(main_name, str(week) + "wPI3.png", "3 курс", course_3_PI_rows)
        new_image = merge_images(str(week) + "main3.png", str(week) + "wPI3.png", str(week) + "wPI3.png")
        timetable_pics[week - 1].append(str(week) + "wPI3.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_3_E_rows != -1):
        excel2img.export_img(main_name, str(week) + "wE3.png", "3 курс", course_3_E_rows)
        new_image = merge_images(str(week) + "main3.png", str(week) + "wE3.png", str(week) + "wE3.png")
        timetable_pics[week - 1].append(str(week) + "wE3.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_3_UB_rows != -1):
        excel2img.export_img(main_name, str(week) + "wUB3.png", "3 курс", course_3_UB_rows)
        new_image = merge_images(str(week) + "main3.png", str(week) + "wUB3.png", str(week) + "wUB3.png")
        timetable_pics[week - 1].append(str(week) + "wUB3.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_3_I_rows != -1):
        excel2img.export_img(main_name, str(week) + "wI3.png", "3 курс", course_3_I_rows)
        new_image = merge_images(str(week) + "main3.png", str(week) + "wI3.png", str(week) + "wI3.png")
        timetable_pics[week - 1].append(str(week) + "wI3.png")
    else:
        timetable_pics[week - 1].append("empty")
    if (course_3_U_rows != -1):
        excel2img.export_img(main_name, str(week) + "wU3.png", "3 курс", course_3_U_rows)
        new_image = merge_images(str(week) + "main3.png", str(week) + "wU3.png", str(week) + "wU3.png")
        timetable_pics[week - 1].append(str(week) + "wU3.png")
    else:
        timetable_pics[week - 1].append("empty")
    excel2img.export_img(main_name, str(week) + "main4.png", "4 курс", rows_time_4)
    if (course_4_I_rows != -1):
        excel2img.export_img(main_name, str(week) + "wI4.png", "4 курс", course_4_I_rows)
        new_image = merge_images(str(week) + "main4.png", str(week) + "wI4.png", str(week) + "wI4.png")
        timetable_pics[week - 1].append(str(week) + "wI4.png")
    else:
        timetable_pics[week - 1].append("empty")


# создаем картинки
if number_of_schedules == 2:
    create_images("this_week.xlsx", 1)
    create_images("next_week.xlsx", 2)
elif number_of_schedules == 1:
    create_images("this_week.xlsx", 1)

print("Картинки скачались")

changes = False
token = "6209739345:AAHw9U9Z-0f5n1Q4Fg__4mGgbE4DCuTSFTI"
bot = telebot.TeleBot(token)
def update():  # функция об отправке уведомления об изменениях
    global changes
    with sqlite3.connect("users.db") as connect:
        connect.row_factory = sqlite3.Row
        cursor = connect.cursor()
        cursor.execute('SELECT user_id FROM users')
    if changes is True:
        for r in cursor:
            bot.send_message(r['user_id'], "Расписание обновилось")
        changes = False
# отслеживаем обновления
def f():
    threading.Timer(3600.0, f).start()
    global number_of_schedules
    global URL
    global shedule_names
    global changes
    global next_week_courses
    global this_week_courses
    changes = False
    new_internalLinks = []
    new_shedule_names = []
    i = 1
    response = requests.get(URL)
    soup = BS(response.content, 'html.parser')
    for a in soup.find_all('a'):
        if ("Расписание занятий (") in str(a.text) and i <= 2:
            i += 1
            new_shedule_names.append(str(a.text))
            new_internalLinks.append(a.get('href'))
    new_number_of_schedules = len(new_internalLinks)

    for i in range(new_number_of_schedules):
        if ('//www.hse.ru') in new_internalLinks[i]:
            new_internalLinks[i] = new_internalLinks[i].replace('//www.hse.ru', 'http://students.perm.hse.ru')
        elif ('www.hse.ru') in new_internalLinks[i]:
            new_internalLinks[i] = new_internalLinks[i].replace('www.hse.ru', 'http://students.perm.hse.ru')
        elif ('http://students.perm.hse.ru') not in new_internalLinks[i]:
            new_internalLinks[i] = 'http://students.perm.hse.ru' + new_internalLinks[i]

    if new_number_of_schedules != 0:
        for i in range(new_number_of_schedules):
            new_shedule_names[i] = new_shedule_names[i].replace("Расписание занятий ", "")
    else:
        number_of_schedules = 0
        changes = True

    if new_number_of_schedules == 2 and number_of_schedules == 2:
        if new_shedule_names[0] != shedule_names[0]:
            shedule_names[0] = new_shedule_names[0]
            save_from_www(new_internalLinks[0], 'this_week.xls')
            this_week = xlrd.open_workbook("this_week.xls", formatting_info=True)
            unmerged_cell('this_week.xls', this_week)
            convert_to_xlxs('this_week')
            this_week = load_workbook('this_week.xlsx')
            for i in this_week_courses:
                cells(i, this_week, 'this_week.xlsx')
            colour('this_week.xlsx')
            create_images("this_week.xlsx", 1)
            changes = True
        if new_shedule_names[1] != shedule_names[1]:
            shedule_names[1] = new_shedule_names[1]
            save_from_www(new_internalLinks[1], 'next_week.xls')
            next_week = xlrd.open_workbook("next_week.xls", formatting_info=True)
            unmerged_cell('next_week.xls', next_week)
            convert_to_xlxs('next_week')
            next_week = load_workbook('next_week.xlsx')
            for i in next_week_courses:
                cells(i, next_week, 'next_week.xlsx')
            colour('next_week.xlsx')
            create_images("next_week.xlsx", 2)
            changes = True
    elif new_number_of_schedules == 1 and number_of_schedules == 1:
        if new_shedule_names[0] != shedule_names[0]:
            shedule_names[0] = new_shedule_names[0]
            save_from_www(new_internalLinks[0], 'this_week.xls')
            this_week = xlrd.open_workbook("this_week.xls", formatting_info=True)
            unmerged_cell('this_week.xls', this_week)
            convert_to_xlxs('this_week')
            this_week = load_workbook('this_week.xlsx')
            for i in this_week_courses:
                cells(i, this_week, 'this_week.xlsx')
            colour('this_week.xlsx')
            create_images("this_week.xlsx", 1)
            changes = True
    elif new_number_of_schedules != 0 and number_of_schedules == 0:
        save_from_www(new_internalLinks[0], 'this_week.xls')
        this_week = xlrd.open_workbook("this_week.xls", formatting_info=True)
        unmerged_cell('this_week.xls', this_week)
        convert_to_xlxs('this_week')
        this_week = load_workbook('this_week.xlsx')
        this_week_courses = [this_week["1 курс"], this_week["2 курс"], this_week["3 курс"], this_week["4 курс"]]
        for i in this_week_courses:
            cells(i, this_week, 'this_week.xlsx')
        colour('this_week.xlsx')
        create_images("this_week.xlsx", 1)
        shedule_names.append(new_shedule_names[0])
        number_of_schedules = new_number_of_schedules
        changes = True
        if new_number_of_schedules == 2:
            save_from_www(new_internalLinks[1], 'next_week.xls')
            next_week = xlrd.open_workbook("next_week.xls", formatting_info=True)
            unmerged_cell('next_week.xls', next_week)
            convert_to_xlxs('next_week')
            next_week = load_workbook('next_week.xlsx')
            next_week_courses = [next_week["1 курс"], next_week["2 курс"], next_week["3 курс"], next_week["4 курс"]]
            for i in next_week_courses:
                cells(i, next_week, 'next_week.xlsx')
            colour('next_week.xlsx')
            create_images("next_week.xlsx", 2)
            shedule_names.append(new_shedule_names[1])
    elif new_number_of_schedules == 2 and number_of_schedules == 1:
        if new_shedule_names[0] != shedule_names[0]:
            shedule_names[0] = new_shedule_names[0]
            save_from_www(new_internalLinks[0], 'this_week.xls')
            this_week = xlrd.open_workbook("this_week.xls", formatting_info=True)
            unmerged_cell('this_week.xls', this_week)
            convert_to_xlxs('this_week')
            this_week = load_workbook('this_week.xlsx')
            for i in this_week_courses:
                cells(i, this_week, 'this_week.xlsx')
            colour('this_week.xlsx')
            create_images("this_week.xlsx", 1)
        save_from_www(new_internalLinks[1], 'next_week.xls')
        next_week = xlrd.open_workbook("next_week.xls", formatting_info=True)
        unmerged_cell('next_week.xls', next_week)
        convert_to_xlxs('next_week')
        next_week = load_workbook('next_week.xlsx')
        next_week_courses = [next_week["1 курс"], next_week["2 курс"], next_week["3 курс"], next_week["4 курс"]]
        for i in next_week_courses:
            cells(i, next_week, 'next_week.xlsx')
        create_images("next_week.xlsx", 2)
        colour('next_week.xlsx')
        shedule_names.append(new_shedule_names[1])
        number_of_schedules = new_number_of_schedules
        changes = True
    elif new_number_of_schedules == 1 and number_of_schedules == 2:
        if new_shedule_names[0] != shedule_names[0]:
            shedule_names[0] = new_shedule_names[0]
            save_from_www(new_internalLinks[0], 'this_week.xls')
            this_week = xlrd.open_workbook("this_week.xls", formatting_info=True)
            unmerged_cell('this_week.xls', this_week)
            convert_to_xlxs('this_week')
            this_week = load_workbook('this_week.xlsx')
            for i in this_week_courses:
                cells(i, this_week, 'this_week.xlsx')
            colour('this_week.xlsx')
            create_images("this_week.xlsx", 1)
        shedule_names.remove(shedule_names[1])
        number_of_schedules = new_number_of_schedules
        changes = True
    update()


f()
print("Изменения чекнулись")

'''////////////бот///////////'''



HELP = ("\n/help - вывести список доступных команд \n"
        "/start - начать бота заново \n"
        "/timetable - отправить расписание \n"
        "/data - вывести выбранные настройки\n"
        "При изменениях в расписании вы получите сообщение 'Расписание обновилось'.\n"
        "Во время обновления расписания бот может отвечать на ваши действия с задержкой (не более 7 минут). \nЕсли у "
        "вас возникнут проблемы, вопросы или советы по улучшению бота, смело пишите @Aloshaaaaaaa или @bbelochkka. Мы "
        "будем рады получить обратную связь.")

"....кейборд для курсов....."
keyboard_inline_course = types.InlineKeyboardMarkup(row_width=2)
for i in range(1, 5):
    keyboard_inline_course.add(types.InlineKeyboardButton(text=str(i) + " курс", callback_data=str(i) + " курс"))

# /////////////////1 курс///////////////////////#

faculties_1st = ["РИС", "МБ", "История", "Юриспруденция", "ИЯ"]

"....кейборд для направлений 1 курса....."
keyboard_inline_1st_Faculties = types.InlineKeyboardMarkup()
for item in faculties_1st:
    keyboard_inline_1st_Faculties.add(types.InlineKeyboardButton(text=item, callback_data=item))

# //////////////////2 курс/////////////////////////#

faculties_2nd = ["БИ", "ПИ", "Экономика", "УБ", "История", "Юриспруденция", "ИЯ"]

"....кейборд для направлений 2 курса....."
keyboard_inline_2nd_Faculties = types.InlineKeyboardMarkup()
for item in faculties_2nd:
    keyboard_inline_2nd_Faculties.add(types.InlineKeyboardButton(text=item, callback_data=item))

# /////////////////3 курс/////////////////////////#

"....кейборд для направлений 3 курса....."
faculties_3d = ["БИ", "ПИ", "Экономика", "УБ", "История", "Юриспруденция"]
keyboard_inline_3d_Faculties = types.InlineKeyboardMarkup()
for item in faculties_3d:
    keyboard_inline_3d_Faculties.add(types.InlineKeyboardButton(text=item, callback_data=item))

# /////////////////4 курс/////////////////////////#

"....кейборд для направлений 4 курса....."
keyboard_inline_4d_Faculties = types.InlineKeyboardMarkup()
keyboard_inline_4d_Faculties.add(types.InlineKeyboardButton(text="История", callback_data="История"))


# ///////////// функции для бд/////////////

def add_user(message):  # добавление юзера в таблицу
    with sqlite3.connect("users.db") as connect:
        cursor = connect.cursor()
        person_id = message.chat.id
        cursor.execute(f"SELECT user_id FROM users WHERE user_id ={person_id}")
        data = cursor.fetchone()
        if data is None:
            user_id = [message.chat.id]
            cursor.execute("INSERT INTO users (user_id) VALUES(?);", user_id)
        else:
            pass


def change_course(data_course, message):  # записываем курс в бд
    with sqlite3.connect("users.db") as connect:
        cursor = connect.cursor()
        cursor.execute('UPDATE users SET course = ? WHERE user_id = ?', (data_course, message.chat.id))


def change_faculty(data_faculty, message):  # записываем направление в бд
    with sqlite3.connect("users.db") as connect:
        cursor = connect.cursor()
        cursor.execute('UPDATE users SET faculty = ? WHERE user_id = ?', (data_faculty, message.chat.id))


def change_number(number, message):  # записываем номер картинки в бд
    with sqlite3.connect("users.db") as connect:
        cursor = connect.cursor()
        cursor.execute('UPDATE users SET number_of_picture = ? WHERE user_id = ?', (number, message.chat.id))


def get_data(message):  # достаем информацию из бд
    with sqlite3.connect("users.db") as connect:
        cursor = connect.cursor()
        tmp = message.chat.id
        cursor.execute('SELECT course, faculty, number_of_picture FROM users WHERE user_id = ?', (int(tmp),))
        rows = cursor.fetchone()
        return rows


def clear(message):  # удаление  бд
    with sqlite3.connect("users.db") as connect:
        cursor = connect.cursor()
        cursor.execute('UPDATE users SET course = ? WHERE user_id = ?', ("0", message.chat.id))
        cursor.execute('UPDATE users SET faculty = ? WHERE user_id = ?', ("направление не выбрано", message.chat.id))
        cursor.execute('UPDATE users SET number_of_picture = ? WHERE user_id = ?', (0, message.chat.id))


'''def update():  # функция об отправке уведомления об изменениях
    global changes
    with sqlite3.connect("users.db") as connect:
        connect.row_factory = sqlite3.Row
        cursor = connect.cursor()
        cursor.execute('SELECT user_id FROM users')
    if changes is True:
        for r in cursor:
            bot.send_message(r['user_id'], "Расписание обновилось")
        changes = False'''




@bot.message_handler(commands=["help"])  # отправка команд бота
def help(message):
    bot.send_message(message.chat.id, HELP)


@bot.message_handler(commands=["course"])  # выбрать курс
def choose_course(message):
    bot.send_message(message.chat.id, "Выберите курс", reply_markup=keyboard_inline_course)


@bot.message_handler(commands=["faculty"])  # выбрать направление
def choose_faculty(message):
    course = get_data(message)
    if course[0] == 1:
        bot.send_message(message.chat.id, "Выберите направление", reply_markup=keyboard_inline_1st_Faculties)
    elif course[0] == 2:
        bot.send_message(message.chat.id, "Выберите направление", reply_markup=keyboard_inline_2nd_Faculties)
    elif course[0] == 3:
        bot.send_message(message.chat.id, "Выберите направление", reply_markup=keyboard_inline_3d_Faculties)
    else:
        bot.send_message(message.chat.id, "Выберите направление", reply_markup=keyboard_inline_4d_Faculties)


@bot.message_handler(commands=["set_photo"])  # каждому пользователю ставим номер картинки
def set_photo(message):
    course = get_data(message)
    faculty = get_data(message)
    if course[0] == 1:
        if faculty[1] == "РИС":
            change_number(0, message)
        elif faculty[1] == "МБ":
            change_number(1, message)
        elif faculty[1] == "История":
            change_number(2, message)
        elif faculty[1] == "Юриспруденция":
            change_number(3, message)
        else:
            change_number(4, message)
    elif course[0] == 2:
        if faculty[1] == "БИ":
            change_number(5, message)
        elif faculty[1] == "ПИ":
            change_number(6, message)
        elif faculty[1] == "Экономика":
            change_number(7, message)
        elif faculty[1] == "УБ":
            change_number(8, message)
        elif faculty[1] == "История":
            change_number(9, message)
        elif faculty[1] == "Юриспруденция":
            change_number(10, message)
        else:
            change_number(11, message)
    elif course[0] == 3:
        if faculty[1] == "БИ":
            change_number(12, message)
        elif faculty[1] == "ПИ":
            change_number(13, message)
        elif faculty[1] == "Экономика":
            change_number(14, message)
        elif faculty[1] == "УБ":
            change_number(15, message)
        elif faculty[1] == "История":
            change_number(16, message)
        else:
            change_number(17, message)
    else:
        change_number(18, message)


@bot.message_handler(commands=["start"]) # запускаем бота
def start(message):
    name = message.from_user.first_name
    text = "Здравствуйте, " + name + "! Я Воронова Шедьюла Экономовна, та самая вышкинская ворона. Я делаю ученическую " \
                                     "жизнь проще, а именно отправляю расписание занятий студентам бакалавриата НИУ ВШЭ " \
                                     "г.Перми. Если возникнут трудности, воспользуйтесь командой /help"
    bot.send_message(message.chat.id, text)
    add_user(message)
    choose_course(message)


@bot.message_handler(commands=["show"]) # получаем выбранные настройки пользователя
def show(message):
    data = get_data(message)
    return str(data[0]) + " курс" + "\n" + "Направление: " + data[1] + "\n"


@bot.message_handler(commands=["timetable"]) # Показать расписание
def timetable(message):
    data = get_data(message)
    if data[0] == 0 or data[1] == "направление не выбрано":
        bot.send_message(message.chat.id, "Для получения расписания сначала выберите настройки")
    elif data[0] == 1 and data[1] not in faculties_1st:
        bot.send_message(message.chat.id, "Это направление отсутствует на выбранном курсе")
    elif data[0] == 2 and data[1] not in faculties_2nd:
        bot.send_message(message.chat.id, "Это направление отсутствует на выбранном курсе")
    elif data[0] == 3 and data[1] not in faculties_3d:
        bot.send_message(message.chat.id, "Это направление отсутствует на выбранном курсе")
    elif data[0] == 4 and data[1] != "История":
        bot.send_message(message.chat.id, "Это направление отсутствует на выбранном курсе")
    else:
        keyboard_inline_timetable = types.InlineKeyboardMarkup(row_width=1)
        if number_of_schedules == 2:
            this_week = types.InlineKeyboardButton(text=shedule_names[0], callback_data="current")
            next_week = types.InlineKeyboardButton(text=shedule_names[1], callback_data="next")
            keyboard_inline_timetable.add(this_week, next_week)
            bot.send_message(message.chat.id, "Выберите расписание", reply_markup=keyboard_inline_timetable)
        elif number_of_schedules == 1:
            this_week = types.InlineKeyboardButton(text=shedule_names[0], callback_data="current")
            keyboard_inline_timetable.add(this_week)
            bot.send_message(message.chat.id, "Выберите расписание", reply_markup=keyboard_inline_timetable)
        else:
            bot.send_message(message.chat.id, "Пар нет, отдыхайте :)")


@bot.message_handler(commands=["photo"]) # отправляем картинку с расписанием
def send_photo(message, week):
    data = get_data(message)
    name = str(timetable_pics[week][data[2]])
    if name == "empty":
        bot.send_message(message.chat.id, "У вас нет пар на этой неделе. Отдыхайте! :)")
    else:
        photo = open(name, "rb")
        bot.send_document(message.chat.id, photo)


@bot.callback_query_handler(func=lambda call: True) # ответ на кнопки
def answer(call):
    if "курс" in call.data:
        change_course(int(call.data[0]), call.message)
        bot.send_message(call.message.chat.id, "Вы выбрали " + call.data)
        choose_faculty(call.message)
    elif call.data in faculties_2nd or call.data in faculties_1st:
        change_faculty(call.data, call.message)
        bot.send_message(call.message.chat.id, "Направление, которое вы выбрали - " + call.data)
        set_photo(call.message)
        keyboard_reply = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item_show = types.KeyboardButton("Показать расписание")
        item_data = types.KeyboardButton("Настройки")
        item_helper = types.KeyboardButton("Помощь")
        keyboard_reply.add(item_show, item_data, item_helper)
        bot.send_message(call.message.chat.id, show(call.message), reply_markup=keyboard_reply)
    elif call.data == "Да":
        clear(call.message)
        choose_course(call.message)
    elif call.data == "current":
        send_photo(call.message, 0)
    elif call.data == "next":
        send_photo(call.message, 1)
    else:
        bot.send_message(call.message.chat.id, "Хорошо, как скажете ;)")


@bot.message_handler(commands=["data"]) # Настройки
def data(message):
    data = get_data(message)
    if data[0] is None:
        add_user(message)
    bot.send_message(message.chat.id, str(data[0]) + " курс" + "\n" + "Направление: " + data[1] + "\n")
    keyboard_inline_data = types.InlineKeyboardMarkup()
    yes = types.InlineKeyboardButton(text="Да", callback_data="Да")
    no = types.InlineKeyboardButton(text="Нет", callback_data="Нет")
    keyboard_inline_data.add(yes, no)
    bot.send_message(message.chat.id, "Хотите поменять настройки?", reply_markup=keyboard_inline_data)


@bot.message_handler(commands=["helper"]) # Помощь
def helper(message):
    bot.send_message(message.chat.id, HELP)


@bot.message_handler(content_types=["text"]) # реакция бота на текст
def hi(message):
    if message.text.lower() == "привет" or message.text.lower() == "приветик" or message.text.lower() == "здравствуйте":
        bot.send_message(message.chat.id, "Приветствуем!")
    elif message.text.lower() == "пока" or message.text.lower() == "прощай" or message.text.lower() == "до свидания" \
            or message.text.lower() == "спасибо":
        bot.send_message(message.chat.id, "Заглядывайте еще!")
    elif message.text == "Настройки":
        data(message)
    elif message.text == "Показать расписание":
        timetable(message)
    elif message.text == "Помощь":
        helper(message)
    else:
        text = "Наш бот пока не может разобрать то, что вы написали..."
        bot.send_message(message.chat.id, text + HELP)
        add_user(message)


if __name__ == "__main__":   # крутим бота постоянно
    while True:
        try:
            bot.polling(none_stop=True)
        except Exception as e:
            time.sleep(3)
            print(e)
